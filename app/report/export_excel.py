#coding:utf-8

import sys,os,re
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border,Side, Alignment, Protection
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.comments import Comment
from .. models import User,Trouble_repo,Trouble_repo_add,Month_trouble_repo,Month_trouble_log,Anomaly_log
from config import basedir

from openpyxl.chart import (
    PieChart,
    ProjectedPieChart,
    Reference
)
from openpyxl.chart.series import DataPoint


def trouble(trouble_times,stab_per,trouble_times_1,stab_per_1,trouble_list,title,head):

    wb = Workbook()
    
    #背景颜色
    redFill = PatternFill(start_color='A52A2A',end_color='A52A2A',fill_type='solid')
    headFill = PatternFill(start_color='357ebd',end_color='357ebd',fill_type='solid')
    
    
    #单元格对齐样式
    headtext = Alignment(horizontal='center',vertical='center')
    bodytext = Alignment(horizontal='center',vertical='center',wrap_text=True)
    
    #字体样式
    ft = Font(color='ffffff',size=14,bold=True)
    ft_black = Font(color='000000',size=14,bold=True)
    
    #边框样式
    border = Border(left=Side(border_style='thin',color='FF000000'),
    right=Side(border_style='thin',color='FF000000'),
    top=Side(border_style='thin',color='FF000000'),
    bottom=Side(border_style='thin',color='FF000000'),
    diagonal=Side(border_style='thin',color='FF000000'),
    diagonal_direction=20,
    outline=Side(border_style='thin',color='FF000000'),
    vertical=Side(border_style='thin',color='FF000000'),
    horizontal=Side(border_style='thin',color='FF000000')
    )
    
    protection = Protection(locked=False,hidden=True)

    
    ws = wb.create_sheet(u'故障报告', 0)
    
    #设置行高
    ws.row_dimensions[1].height = 40.0
    ws.row_dimensions[2].height = 40.0

    #ws['A1']=head
    
    
    ws['B1']= u'核心服务故障时间:%s分钟 ' % trouble_times
    ws['C1']= u'稳定性:%s%%' % stab_per
    ws['F1']=u'非核心服务故障时间:%s分钟 ' % trouble_times_1
    ws['G1']=u'稳定性:%s%%' % stab_per_1

    head_text = [u"日期",u"运营中心",u"业务模块",u"事件",u"影响范围",u"是否内部故障",u"影响时长(分钟)",u"是否影响用户体验",u"影响用户",u"直接经济损失(美元)",u"数据来源",u"是否核心服务",u"故障类型",u"处理负责人",u"归属",u"状态",u"故障原因",u"处理过程",u"教训总结",u"改进"]

    ws.append(head_text)

    ws['C2'].comment = Comment(text="游戏明\n储值\n登陆\n后台\n所有", author="业务模块")
    ws['D2'].comment = Comment(text="描素事件现象", author="事件")
    ws['F2'].comment = Comment(text="如网络、CDN、原厂、渠道等外部因素引起的故障都不是内部故障", author="是否内部故障")
    ws['I2'].comment = Comment(text="1、昨天在线数据-今天储值数据；\n2、如数据值为负，表示故障时段数据比昨天上升，则故障影响不大；\n3、若数据为0，则表示无影响。", author="影响用户")
    ws['J2'].comment = Comment(text="1、昨天储值数据-今天储值数据；\n2、如数据值为负，表示故障时段数据比昨天上升，则故障影响不大 \n3、若数据为0，则表示无影响", author="经济损失")
    

    n = 3
    for trouble in trouble_list:
    
        ws.append(trouble)
        for i in 'ABCDEFGHIJKLMNOPQRST':
            ws['%s%d' % (i,n) ].border = border
            ws['%s%d' % (i,n) ].alignment = bodytext
            ws['%s%d' % (i,n) ].protection = protection
        ws['Q%d' % n].alignment = Alignment(horizontal='general', vertical='center', wrap_text=True)
        ws['R%d' % n].alignment = Alignment(horizontal='general',vertical='center',wrap_text=True)
        ws['S%d' % n].alignment = Alignment(horizontal='general',vertical='center',wrap_text=True)
        ws['T%d' % n].alignment = Alignment(horizontal='general',vertical='center',wrap_text=True)
        ws.row_dimensions[n].height = 30.0
        n+=1
        
        
    for i in 'ABCDEFGHIJKLMNOPQRST':
        #ws['%s1' % i ].fill = redFill
        ws['%s2' % i ].fill = headFill

        ws['%s2' % i ].border = border

        ws['%s1' % i ].font = ft_black
        ws['%s2' % i ].font = ft
        
        ws['%s1' % i ].alignment = headtext
        ws['%s2' % i ].alignment = headtext
        
        ws['%s1' % i ].protection = protection

        
        #设置宽度

        ws.column_dimensions[i].width = 20.0

    ws.column_dimensions['B'].width = 28.0
    ws.column_dimensions['F'].width = 30.0
    ws.column_dimensions['Q'].width = 35.0
    ws.column_dimensions['R'].width = 40.0


  #  ws['F2'].fill = PatternFill(start_color='e26b0a',end_color='e26b0a',fill_type='solid')
  #  ws['G2'].fill = PatternFill(start_color='e26b0a',end_color='e26b0a',fill_type='solid')

    
    wb.save(title)


def anomaly(anomaly_list, title):
    wb = Workbook()

    # 背景颜色
    redFill = PatternFill(start_color='A52A2A', end_color='A52A2A', fill_type='solid')
    headFill = PatternFill(start_color='357ebd', end_color='357ebd', fill_type='solid')

    # 单元格对齐样式
    headtext = Alignment(horizontal='center', vertical='center')
    bodytext = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 字体样式
    ft = Font(color='ffffff', size=14, bold=True)

    # 边框样式
    border = Border(left=Side(border_style='thin', color='FF000000'),
                    right=Side(border_style='thin', color='FF000000'),
                    top=Side(border_style='thin', color='FF000000'),
                    bottom=Side(border_style='thin', color='FF000000'),
                    diagonal=Side(border_style='thin', color='FF000000'),
                    diagonal_direction=20,
                    outline=Side(border_style='thin', color='FF000000'),
                    vertical=Side(border_style='thin', color='FF000000'),
                    horizontal=Side(border_style='thin', color='FF000000')
                    )

    protection = Protection(locked=False, hidden=True)

    ws = wb.create_sheet(u'异常记录', 0)

    # 设置行高
    ws.row_dimensions[1].height = 35.0

    head_text = [u"异常事件", u"运营中心", u"异常反馈来源",u"异常反馈类型",u"业务模块",u"异常级别", u"是否误报", u"是否回收/维护", u"影响范围", u"发生时间", u"报错时间",
                 u"开始处理时间", u"处理结束时间",u"异常归属", u"处理人", u"处理结果", u"5分钟反馈", u"15分钟反馈", u"30分钟反馈", u"1小时反馈", u"2小时反馈",
                 u"监控评价", u"监控跟进人"]

    ws.append(head_text)

    n = 2
    for trouble in anomaly_list:
        ws.row_dimensions[n].height = 25.0
        ws.append(trouble)

        for i in 'ABCDEFGHIJKLMNOPQRSTUVW':
            ws['%s%d' % (i, n)].border = border
            ws['%s%d' % (i, n)].alignment = bodytext
            ws['%s%d' % (i, n)].protection = protection
        ws['P%d' % n].alignment = Alignment(horizontal='general', vertical='center', wrap_text=True)
        ws['Q%d' % n].alignment = Alignment(horizontal='general', vertical='center', wrap_text=True)
        ws['R%d' % n].alignment = Alignment(horizontal='general', vertical='center', wrap_text=True)
        ws['S%d' % n].alignment = Alignment(horizontal='general', vertical='center', wrap_text=True)
        ws['T%d' % n].alignment = Alignment(horizontal='general', vertical='center', wrap_text=True)
        ws['U%d' % n].alignment = Alignment(horizontal='general', vertical='center', wrap_text=True)

        n += 1

    for i in 'ABCDEFGHIJKLMNOPQRSTUVW':
        ws['%s1' % i].fill = headFill

        ws['%s1' % i].border = border

        ws['%s1' % i].font = ft

        ws['%s1' % i].alignment = headtext

        ws['%s1' % i].protection = protection

        # 设置宽度
        ws.column_dimensions[i].width = 20.0

    wb.save(title)


def monthrepo(**data):
    wb = Workbook()

    # 背景颜色
    titleFill = PatternFill(start_color='A52A2A', end_color='A52A2A', fill_type='solid')
    headFill = PatternFill(start_color='357ebd', end_color='357ebd', fill_type='solid')

    # 单元格对齐样式
    headtext = Alignment(horizontal='center', vertical='center')
    bodytext = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 字体样式
    ft = Font(name=u'微软雅黑',color='ffffff', size=14, bold=False)

    ft_ps = Font(color='FF0000', size=12, bold=True)

    #ft_pm = Font(color='e26b0a', size=12, bold=True)

    ft_pm = Font(name=u'微软雅黑',color='FFC000', size=12, bold=True)

    # 边框样式
    border = Border(left=Side(border_style='thin', color='FF000000'),
                    right=Side(border_style='thin', color='FF000000'),
                    top=Side(border_style='thin', color='FF000000'),
                    bottom=Side(border_style='thin', color='FF000000'),
                    diagonal=Side(border_style='thin', color='FF000000'),
                    diagonal_direction=20,
                    outline=Side(border_style='thin', color='FF000000'),
                    vertical=Side(border_style='thin', color='FF000000'),
                    horizontal=Side(border_style='thin', color='FF000000')
                    )

    protection = Protection(locked=False, hidden=True)

    ws = wb.create_sheet(u'月故障分析', 0)

    # 设置行高

    ws.row_dimensions[1].height = 35.0


    ###############################表格一###################################
    ws.merge_cells('A1:E1')
    ws.merge_cells('A2:E2')
    ws['A1'] = u'各类型故障信息统计'
    ws['A2'] = u'月度可用率基准值暂定为99%，各项指标可用率均超过99%，处于可接受范围'
    ws['A2'].font = ft_ps


    thead_1 = [u'故障指标',u'运营中心',u'业务模块',u'故障时间',u'月可用率']
    ws.append(thead_1)


    AE_N =data['AE_row']
    AE_N += 3
    ws.merge_cells('B4:B%d' % AE_N)
    ws['B4'] = u'亚欧'


    HT_N = data['HT_row']
    HT_N += AE_N
    AE_N += 1

    ws.merge_cells('B%d:B%d' % (AE_N,HT_N))
    ws['B%d' % AE_N] = u'港台'

    KR_N = data['KR_row']
    KR_N += HT_N
    HT_N += 1

    ws.merge_cells('B%d:B%d' % (HT_N,KR_N))
    ws['B%d' % HT_N] = u'韩国'

    CN_N = data['CN_row']
    CN_N += KR_N
    KR_N +=1

    ws.merge_cells('B%d:B%d' % (KR_N,CN_N))
    ws['B%d' % KR_N] = u'国内'

    GB_N = data['GB_row']
    GB_N += CN_N
    CN_N +=1

    ws.merge_cells('B%d:B%d' % (CN_N,GB_N))
    ws['B%d' % CN_N] = u'全球'

    ALL_N = data['ALL_row']
    if ALL_N >0:
        ALL_N += GB_N
        GB_N+=1
        ws.merge_cells('B%d:B%d' % (GB_N,ALL_N))
        ws['B%d' % GB_N] = u'所有地区'

    else:
        pass

    #####################################################
    N = 4
    #亚欧一级指标
    ws['C%d' % N] = u'登陆'
    ws['D%d' % N] = data['trouble_time_AE_login']
    ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_AE_login']) / data['month_time']) * 100)

    N += 1
    ws['C%d' % N] = u'储值'
    ws['D%d' % N] = data['trouble_time_AE_store']
    ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_AE_store']) / data['month_time']) * 100)

    N += 1
    ws['C%d' % N] = u'注册'
    ws['D%d' % N] = data['trouble_time_AE_register']
    ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_AE_register']) / data['month_time']) * 100)

    N += 1
    ws['C%d' % N] = u'游戏故障'
    ws['D%d' % N] = data['trouble_time_AE_game']
    ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_AE_game']) / data['month_time']) * 100)

    if data['trouble_time_AE_all'] >0:
        N +=1
        ws['C%d' % N] = u'ALL'
        ws['D%d' % N] = data['trouble_time_AE_all']
        ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_AE_all']) / data['month_time']) * 100)

    #####################################################
    # 港台一级指标
    N += 1
    ws['C%d' % N] = u'登陆'
    ws['D%d' % N] = data['trouble_time_HT_login']
    ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_HT_login']) / data['month_time']) * 100)

    N += 1
    ws['C%d' % N] = u'储值'
    ws['D%d' % N] = data['trouble_time_HT_store']
    ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_HT_store']) / data['month_time']) * 100)

    N += 1
    ws['C%d' % N] = u'注册'
    ws['D%d' % N] = data['trouble_time_HT_register']
    ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_HT_register']) / data['month_time']) * 100)

    N += 1
    ws['C%d' % N] = u'游戏故障'
    ws['D%d' % N] = data['trouble_time_HT_game']
    ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_HT_game']) / data['month_time']) * 100)

    if data['trouble_time_HT_all'] > 0:
        N += 1
        ws['C%d' % N] = u'ALL'
        ws['D%d' % N] = data['trouble_time_HT_all']
        ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_HT_all']) / data['month_time']) * 100)


    #####################################################
    # 韩国一级指标
    N += 1
    ws['C%d' % N] = u'登陆'
    ws['D%d' % N] = data['trouble_time_KR_login']
    ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_KR_login']) / data['month_time']) * 100)

    N += 1
    ws['C%d' % N] = u'储值'
    ws['D%d' % N] = data['trouble_time_KR_store']
    ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_KR_store']) / data['month_time']) * 100)

    N += 1
    ws['C%d' % N] = u'注册'
    ws['D%d' % N] = data['trouble_time_KR_register']
    ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_KR_register']) / data['month_time']) * 100)

    N += 1
    ws['C%d' % N] = u'游戏故障'
    ws['D%d' % N] = data['trouble_time_KR_game']
    ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_KR_game']) / data['month_time']) * 100)

    if data['trouble_time_KR_all'] > 0:
        N += 1
        ws['C%d' % N] = u'ALL'
        ws['D%d' % N] = data['trouble_time_KR_all']
        ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_KR_all']) / data['month_time']) * 100)

    #####################################################
    # 国内一级指标
    N += 1
    ws['C%d' % N] = u'登陆'
    ws['D%d' % N] = data['trouble_time_CN_login']
    ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_CN_login']) / data['month_time']) * 100)

    N += 1
    ws['C%d' % N] = u'储值'
    ws['D%d' % N] = data['trouble_time_CN_store']
    ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_CN_store']) / data['month_time']) * 100)

    N += 1
    ws['C%d' % N] = u'注册'
    ws['D%d' % N] = data['trouble_time_CN_register']
    ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_CN_register']) / data['month_time']) * 100)

    N += 1
    ws['C%d' % N] = u'游戏故障'
    ws['D%d' % N] = data['trouble_time_CN_game']
    ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_CN_game']) / data['month_time']) * 100)

    if data['trouble_time_CN_all'] > 0:
        N += 1
        ws['C%d' % N] = u'ALL'
        ws['D%d' % N] = data['trouble_time_CN_all']
        ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_CN_all']) / data['month_time']) * 100)

    #####################################################
    # 全球一级指标
    N += 1
    ws['C%d' % N] = u'登陆'
    ws['D%d' % N] = data['trouble_time_GB_login']
    ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_GB_login']) / data['month_time']) * 100)

    N += 1
    ws['C%d' % N] = u'储值'
    ws['D%d' % N] = data['trouble_time_GB_store']
    ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_GB_store']) / data['month_time']) * 100)

    N += 1
    ws['C%d' % N] = u'注册'
    ws['D%d' % N] = data['trouble_time_GB_register']
    ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_GB_register']) / data['month_time']) * 100)

    N += 1
    ws['C%d' % N] = u'游戏故障'
    ws['D%d' % N] = data['trouble_time_GB_game']
    ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_GB_game']) / data['month_time']) * 100)

    if data['trouble_time_GB_all'] > 0:
        N += 1
        ws['C%d' % N] = u'ALL'
        ws['D%d' % N] = data['trouble_time_GB_all']
        ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_GB_all']) / data['month_time']) * 100)


    if ALL_N >0:
        if data['trouble_time_ALL_login'] >0:
            N += 1
            ws['C%d' % N] = u'登陆'
            ws['D%d' % N] = data['trouble_time_ALL_login']
            ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_ALL_login']) / data['month_time']) * 100)
        if data['trouble_time_ALL_store'] > 0:
            N += 1
            ws['C%d' % N] = u'储值'
            ws['D%d' % N] = data['trouble_time_ALL_store']
            ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_ALL_store']) / data['month_time']) * 100)
        if data['trouble_time_ALL_register'] > 0:
            N += 1
            ws['C%d' % N] = u'注册'
            ws['D%d' % N] = data['trouble_time_ALL_register']
            ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_ALL_register']) / data['month_time']) * 100)
        if data['trouble_time_ALL_game'] > 0:
            N += 1
            ws['C%d' % N] = u'游戏故障'
            ws['D%d' % N] = data['trouble_time_ALL_game']
            ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_ALL_game']) / data['month_time']) * 100)

        if data['trouble_time_ALL_all'] > 0:
            N += 1
            ws['C%d' % N] = u'ALL'
            ws['D%d' % N] = data['trouble_time_ALL_all']
            ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_ALL_all']) / data['month_time']) * 100)

    N += 1
    ws.merge_cells('A4:A%d' % N)
    ws['A4'] = u'一级指标'

    ws.merge_cells('B%d:C%d' % (N,N))

    ws['B%d' % N] = u'合计（一级指标）'
    ws['D%d' % N] = data['trouble_time_is_core']
    ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_is_core']) / data['month_time']) * 100)

    ws['D%d' % N].font = ft_pm
    ws['E%d' % N].font = ft_pm



    #################################
    two_N = N + 1
    N += 1
    if data['list_AE'] and data['list_AE'][0][1] >0:
        AE_n = N
        for i in data['list_AE']:
            if i[1]>0:
                ws['C%d' % N] = i[0]
                ws['D%d' % N] = i[1]
                ws['E%d' % N] = "%.2f%%" % ((1 - float(i[1]) / data['month_time']) * 100)
                N += 1

        ws.merge_cells('B%d:B%d' % (AE_n,(N-1)))
        ws['B%d' % AE_n] = u'亚欧'

    if data['list_HT'] and data['list_HT'][0][1] > 0:

        HT_n = N

        for i in data['list_HT']:
            if i[1] >0:

                ws['C%d' % N] = i[0]
                ws['D%d' % N] = i[1]
                ws['E%d' % N] = "%.2f%%" % ((1 - float(i[1]) / data['month_time']) * 100)
                N += 1


        ws.merge_cells('B%d:B%d' % (HT_n,(N-1)))
        ws['B%d' % HT_n] = u'港台'

    if data['list_KR'] and data['list_KR'][0][1] > 0:

        KR_n = N

        for i in data['list_KR']:
            if i[1] > 0:

                ws['C%d' % N] = i[0]
                ws['D%d' % N] = i[1]
                ws['E%d' % N] = "%.2f%%" % ((1 - float(i[1]) / data['month_time']) * 100)
                N += 1


        ws.merge_cells('B%d:B%d' % (KR_n,(N-1)))
        ws['B%d' % KR_n] = u'韩国'

    if data['list_CN'] and data['list_CN'][0][1] > 0:

        CN_n = N

        for i in data['list_CN']:
            if i[1] >0:

                ws['C%d' % N] = i[0]
                ws['D%d' % N] = i[1]
                ws['E%d' % N] = "%.2f%%" % ((1 - float(i[1]) / data['month_time']) * 100)
                N += 1


        ws.merge_cells('B%d:B%d' % (CN_n,(N-1)))
        ws['B%d' % CN_n] = u'国内'

    if data['list_GB'] and data['list_GB'][0][1] > 0:
        GB_n = N
        for i in data['list_GB']:
            if i[1] >0:
                ws['C%d' % N] = i[0]
                ws['D%d' % N] = i[1]
                ws['E%d' % N] = "%.2f%%" % ((1 - float(i[1]) / data['month_time']) * 100)
                N += 1

        ws.merge_cells('B%d:B%d' % (GB_n,(N-1)))
        ws['B%d' % GB_n] = u'全球'



    ws.merge_cells('B%d:C%d' % (N,N))
    ws['B%d' % N]= u'合计（二级指标）'
    ws['D%d' % N] = data['trouble_time_not_core']
    ws['E%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_not_core']) / data['month_time']) * 100)

    ws.merge_cells('A%d:A%d' % (two_N,N))
    ws['A%d' % two_N] = u'二级指标'

    N += 1
    ws.merge_cells('A%d:C%d' % (N, N))
    ws['A%d' % N] = u'合计（总）'
    ws['D%d' % N] = data['trouble_time_not_core'] + data['trouble_time_is_core']
    ws['E%d' % N] = "%.2f%%" % ((1 - float((data['trouble_time_not_core'] + data['trouble_time_is_core'])) / data['month_time']) * 100)


    for i in "ABCDE":
        ws.column_dimensions[i].width = 20.0
        ws['%s1' % i].fill = titleFill
        ws['%s1' % i].alignment = headtext
        ws['%s1' % i].protection = protection
        ws['%s1' % i].border = border
        ws['%s2' % i].border = border
        ws['%s1' % i].font = ft

        for j in range(3,N+1):
            ws['%s%d' % (i, j)].fill = headFill
            ws['%s%d' % (i, j)].alignment = headtext
            ws['%s%d' % (i, j)].protection = protection
            ws['%s%d' % (i, j)].border = border
            ws['%s%d' % (i, j)].font = ft

    ws['A2'].font = ft_ps






    ###############################表格二###################################

    last_month = Month_trouble_log.query.filter_by(trouble_month=data['last_month_date']).first()

    ws.merge_cells('G1:K1')
    ws['G1'] = u'本月与上月故障信息对比'
    ws['G2'] = u'业务指标'
    ws['H2'] = u'运营中心'
    ws['I2'] = u'业务模块'
    ws['J2'] = u'本月'
    ws['K2'] = u'上月'

    N = 3
    n = 4
    if data['trouble_time_AE_all'] > 0 or last_month.trouble_time_AE_all_core > 0:
        n = 5
    AE = 3
    N += n
    ws.merge_cells('H%d:H%d' % (AE,(N-1)))
    ws['H%d' % AE] = u'亚欧'


    n = 4
    if data['trouble_time_HT_all'] > 0 or last_month.trouble_time_HT_all_core > 0:
        n = 5
    HT = N
    N += n
    ws.merge_cells('H%d:H%d' % (HT,(N-1)))
    ws['H%d' % HT] = u'港台'

    n = 4
    if data['trouble_time_KR_all'] > 0 or last_month.trouble_time_KR_all_core > 0:
        n = 5
    KR = N
    N += n
    ws.merge_cells('H%d:H%d' % (KR,(N-1)))
    ws['H%d' % KR] = u'韩国'


    n = 4
    if data['trouble_time_CN_all'] > 0 or last_month.trouble_time_CN_all_core > 0:
        n = 5
    CN = N
    N += n
    ws.merge_cells('H%d:H%d' % (CN, (N-1)))
    ws['H%d' % CN] = u'国内'


    n = 4
    if data['trouble_time_GB_all'] > 0 or last_month.trouble_time_GB_all_core > 0:
        n = 5
    GB = N
    N += n
    ws.merge_cells('H%d:H%d' % (GB, (N-1)))
    ws['H%d' % GB] = u'全球'

    if data['ALL_row_pk'] > 0:
        n = data['ALL_row_pk']
        ALL = N
        N += n
        ws.merge_cells('H%d:H%d' % (ALL, (N-1)))
        ws['H%d' % ALL] = u'所有地区'



    ws.merge_cells('H%d:I%d' % (N,N))
    N += 1
    ws.merge_cells('H%d:I%d' % (N,N))
    ws.merge_cells('G3:G%d' % N)
    ws['G3'] = u'核心业务'



    N = 3
    #亚欧一级指标
    ws['I%d' % N] = u'登陆'
    ws['J%d' % N] = data['trouble_time_AE_login']
    ws['K%d' % N] = last_month.trouble_time_AE_login_core

    N += 1
    ws['I%d' % N] = u'储值'
    ws['J%d' % N] = data['trouble_time_AE_store']
    ws['K%d' % N] = last_month.trouble_time_AE_store_core

    N += 1
    ws['I%d' % N] = u'注册'
    ws['J%d' % N] = data['trouble_time_AE_register']
    ws['K%d' % N] = last_month.trouble_time_AE_register_core

    N += 1
    ws['I%d' % N] = u'游戏故障'
    ws['J%d' % N] = data['trouble_time_AE_game']
    ws['K%d' % N] = last_month.trouble_time_AE_game_core

    if data['trouble_time_AE_all'] > 0 or last_month.trouble_time_AE_all_core > 0:
        N += 1
        ws['I%d' % N] = u'ALL'
        ws['J%d' % N] = data['trouble_time_AE_all']
        ws['K%d' % N] = last_month.trouble_time_AE_all_core


    # 港台一级指标
    N += 1
    ws['I%d' % N] = u'登陆'
    ws['J%d' % N] = data['trouble_time_HT_login']
    ws['K%d' % N] = last_month.trouble_time_HT_login_core

    N += 1
    ws['I%d' % N] = u'储值'
    ws['J%d' % N] = data['trouble_time_HT_store']
    ws['K%d' % N] = last_month.trouble_time_HT_store_core

    N += 1
    ws['I%d' % N] = u'注册'
    ws['J%d' % N] = data['trouble_time_HT_register']
    ws['K%d' % N] = last_month.trouble_time_HT_register_core

    N += 1
    ws['I%d' % N] = u'游戏故障'
    ws['J%d' % N] = data['trouble_time_HT_game']
    ws['K%d' % N] = last_month.trouble_time_HT_game_core

    if data['trouble_time_HT_all'] > 0 or last_month.trouble_time_HT_all_core > 0:
        N += 1
        ws['I%d' % N] = u'ALL'
        ws['J%d' % N] = data['trouble_time_HT_all']
        ws['K%d' % N] = last_month.trouble_time_HT_all_core



    # 韩国一级指标
    N += 1
    ws['I%d' % N] = u'登陆'
    ws['J%d' % N] = data['trouble_time_KR_login']
    ws['K%d' % N] = last_month.trouble_time_KR_login_core
    N += 1
    ws['I%d' % N] = u'储值'
    ws['J%d' % N] = data['trouble_time_KR_store']
    ws['K%d' % N] = last_month.trouble_time_KR_store_core
    N += 1
    ws['I%d' % N] = u'注册'
    ws['J%d' % N] = data['trouble_time_KR_register']
    ws['K%d' % N] = last_month.trouble_time_KR_register_core
    N += 1
    ws['I%d' % N] = u'游戏故障'
    ws['J%d' % N] = data['trouble_time_KR_game']
    ws['K%d' % N] = last_month.trouble_time_KR_game_core

    if data['trouble_time_KR_all'] > 0 or last_month.trouble_time_KR_all_core > 0:
        N += 1
        ws['I%d' % N] = u'ALL'
        ws['J%d' % N] = data['trouble_time_KR_all']
        ws['K%d' % N] = last_month.trouble_time_KR_all_core

    # 国内一级指标
    N += 1
    ws['I%d' % N] = u'登陆'
    ws['J%d' % N] = data['trouble_time_CN_login']
    ws['K%d' % N] = last_month.trouble_time_CN_login_core
    N += 1
    ws['I%d' % N] = u'储值'
    ws['J%d' % N] = data['trouble_time_CN_store']
    ws['K16'] = last_month.trouble_time_CN_store_core
    N += 1
    ws['I%d' % N] = u'注册'
    ws['J%d' % N] = data['trouble_time_CN_register']
    ws['K%d' % N] = last_month.trouble_time_CN_register_core
    N += 1
    ws['I%d' % N] = u'游戏故障'
    ws['J%d' % N] = data['trouble_time_CN_game']
    ws['K%d' % N] = last_month.trouble_time_CN_game_core

    if data['trouble_time_CN_all'] > 0 or last_month.trouble_time_CN_all_core > 0:
        N += 1
        ws['I%d' % N] = u'ALL'
        ws['J%d' % N] = data['trouble_time_CN_all']
        ws['K%d' % N] = last_month.trouble_time_CN_all_core

    # 全球一级指标
    N += 1
    ws['I%d' % N] = u'登陆'
    ws['J%d' % N] = data['trouble_time_GB_login']
    ws['K%d' % N] = last_month.trouble_time_GB_login_core
    N += 1
    ws['I%d' % N] = u'储值'
    ws['J%d' % N] = data['trouble_time_GB_store']
    ws['K%d' % N] = last_month.trouble_time_GB_store_core
    N += 1
    ws['I%d' % N] = u'注册'
    ws['J%d' % N] = data['trouble_time_GB_register']
    ws['K%d' % N] = last_month.trouble_time_GB_register_core
    N += 1
    ws['I%d' % N] = u'游戏故障'
    ws['J%d' % N] = data['trouble_time_GB_game']
    ws['K%d' % N] = last_month.trouble_time_GB_game_core

    if data['trouble_time_GB_all'] > 0 or last_month.trouble_time_GB_all_core > 0:
        N += 1
        ws['I%d' % N] = u'ALL'
        ws['J%d' % N] = data['trouble_time_GB_all']
        ws['K%d' % N] = last_month.trouble_time_GB_all_core



    if data['ALL_row_pk'] >0:
        if data['trouble_time_ALL_login'] > 0 or last_month.trouble_time_ALL_login_core > 0:
            N += 1
            ws['I%d' % N] = u'登陆'
            ws['J%d' % N] = data['trouble_time_ALL_login']
            ws['K%d' % N] = last_month.trouble_time_ALL_login_core
        if data['trouble_time_ALL_store'] > 0 or last_month.trouble_time_ALL_store_core > 0:
            N += 1
            ws['I%d' % N] = u'储值'
            ws['J%d' % N] = data['trouble_time_ALL_store']
            ws['K%d' % N] = last_month.trouble_time_ALL_store_core
        if data['trouble_time_ALL_register'] > 0 or last_month.trouble_time_ALL_register_core > 0:
            N += 1
            ws['I%d' % N] = u'注册'
            ws['J%d' % N] = data['trouble_time_ALL_register']
            ws['K%d' % N] = last_month.trouble_time_ALL_register_core
        if data['trouble_time_ALL_game'] > 0 or last_month.trouble_time_ALL_game_core > 0:
            N += 1
            ws['I%d' % N] = u'游戏故障'
            ws['J%d' % N] = data['trouble_time_ALL_game']
            ws['K%d' % N] = last_month.trouble_time_ALL_game_core

        if data['trouble_time_ALL_all'] > 0 or last_month.trouble_time_ALL_all_core > 0:
            N += 1
            ws['I%d' % N] = u'ALL'
            ws['J%d' % N] = data['trouble_time_ALL_all']
            ws['K%d' % N] = last_month.trouble_time_ALL_all_core



    N += 1
    ws['H%d' % N] = u'合计（核心业务）'
    ws['J%d' % N] = data['trouble_time_is_core']
    ws['K%d' % N] = last_month.trouble_time_is_core


    N += 1
    ws['H%d' % N] = u'时长占比'
    ws['J%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_is_core']) / data['month_time']) * 100)
    ws['K%d' % N] = "%.2f%%" % ((1 - float(last_month.trouble_time_is_core) / data['month_time']) * 100)


    two_n = N +1
    N += 1
    if data['list_AE']:
        AE_n = N
        for i in data['list_AE']:
            ws['I%d' % N] = i[0]
            ws['J%d' % N] = i[1]
            ws['K%d' % N] = i[2]
            N += 1
        ws.merge_cells('H%d:H%d' % (AE_n,(N-1)))
        ws['H%d' % AE_n] = u'亚欧'

    if data['list_HT']:
        HT_n = N
        for i in data['list_HT']:
            ws['I%d' % N] = i[0]
            ws['J%d' % N] = i[1]
            ws['K%d' % N] = i[2]
            N += 1

        ws.merge_cells('H%d:H%d' % (HT_n,(N-1)))
        ws['H%d' % HT_n] = u'港台'

    if data['list_KR']:
        KR_n = N
        for i in data['list_KR']:
            ws['I%d' % N] = i[0]
            ws['J%d' % N] = i[1]
            ws['K%d' % N] = i[2]
            N += 1
        ws.merge_cells('H%d:H%d' % (KR_n,(N-1)))
        ws['H%d' % KR_n] = u'韩国'

    if data['list_CN']:

        CN_n = N
        for i in data['list_CN']:
            ws['I%d' % N] = i[0]
            ws['J%d' % N] = i[1]
            ws['K%d' % N] = i[2]
            N += 1
        ws.merge_cells('H%d:H%d' % (CN_n,(N-1)))
        ws['H%d' % CN_n] = u'国内'

    if data['list_GB']:

        GB_n = N
        for i in data['list_GB']:
            ws['I%d' % N] = i[0]
            ws['J%d' % N] = i[1]
            ws['K%d' % N] = i[2]
            N += 1
        ws.merge_cells('H%d:H%d' % (GB_n,(N-1)))
        ws['H%d' % GB_n] = u'全球'

    ws.merge_cells('H%d:I%d' % (N,N))
    ws['H%d' % N]= u'合计（非核心业务）'
    ws['J%d' % N] = data['trouble_time_not_core']
    ws['K%d' % N] = last_month.trouble_time_not_core


    N += 1
    ws.merge_cells('H%d:I%d' % (N,N))
    ws['H%d' % N] = u'时长占比'
    ws['J%d' % N] = "%.2f%%" % ((1 - float(data['trouble_time_not_core']) / data['month_time']) * 100)
    ws['K%d' % N] = "%.2f%%" % ((1 - float(last_month.trouble_time_not_core) / data['month_time']) * 100)


    ws.merge_cells('G%d:G%d' % (two_n, N))
    ws['G%d' % two_n] = '非核心业务'

    N += 1
    ws.merge_cells('G%d:I%d' % (N,N))
    ws['G%d' % N] = u'合计(总)'
    ws['J%d' % N] = data['trouble_time_is_core'] + data['trouble_time_not_core']
    ws['K%d' % N] = int(last_month.trouble_time_is_core) + int(last_month.trouble_time_not_core)


    N += 1
    ws.merge_cells('G%d:I%d' % (N, N))
    ws['G%d' % N] = u'总时长占比'
    ws['J%d' % N] = "%.2f%%" % ((1 - float((data['trouble_time_is_core'] + data['trouble_time_not_core'])) / data['month_time']) * 100)
    ws['K%d' % N] = "%.2f%%" % ((1 - float((int(last_month.trouble_time_is_core) + int(last_month.trouble_time_not_core))) / data['month_time']) * 100)


    for i in "GHIJK":
        ws.column_dimensions[i].width = 20.0
        ws['%s1' % i].fill = titleFill
        ws['%s1' % i].alignment = headtext
        ws['%s1' % i].protection = protection
        ws['%s1' % i].border = border
        ws['%s2' % i].border = border
        ws['%s3' % i].border = border
        ws['%s1' % i].font = ft

        for n in range(2,N+1):
            ws['%s%d' % (i,n)].fill = headFill
            ws['%s%d' % (i, n)].alignment = headtext
            ws['%s%d' % (i, n)].protection = protection
            ws['%s%d' % (i, n)].border = border
            ws['%s%d' % (i, n)].font = ft




    ###############################表格三###################################


    inner_N= N + 7

    ws.merge_cells('A%d:K%d' % (inner_N, inner_N))
    ws['A%d' % inner_N] = '内外故障信息统计'


    ws.merge_cells('A%d:B%d' % (inner_N+1, inner_N+1))
    ws['A%d' % (inner_N +1)] = '归属'
    ws['C%d' % (inner_N + 1)] = '故障时间'
    ws['D%d' % (inner_N + 1)] = '占比'

    ws.merge_cells('A%d:A%d' % (inner_N + 2, inner_N + 4))
    ws['A%d' % (inner_N + 2)] = '内部故障'
    ws['B%d' % (inner_N + 2)] = '运维'
    ws['C%d' % (inner_N + 2)] = data['trouble_time_yw_inner']
    ws['D%d' % (inner_N + 2)] = "%.2f%%" % ((float(data['trouble_time_yw_inner']) / data['trouble_time_all']) * 100)


    ws['B%d' % (inner_N + 3)] = '业务开发'
    ws['C%d' % (inner_N + 3)] = data['trouble_time_ywkf_inner']
    ws['D%d' % (inner_N + 3)] = "%.2f%%" % ((float(data['trouble_time_ywkf_inner']) / data['trouble_time_all']) * 100)

    ws['B%d' % (inner_N + 4)] = '基础开发'
    ws['C%d' % (inner_N + 4)] = data['trouble_time_jckf_inner']
    ws['D%d' % (inner_N + 4)] = "%.2f%%" % ((float(data['trouble_time_jckf_inner']) / data['trouble_time_all']) * 100)

    ws['B%d' % (inner_N + 5)] = '运营'
    ws['C%d' % (inner_N + 5)] = data['trouble_time_yy']
    ws['D%d' % (inner_N + 5)] = "%.2f%%" % ((float(data['trouble_time_yy']) / data['trouble_time_all']) * 100)


    ws['A%d' % (inner_N + 6)] = '外部故障'
    ws['B%d' % (inner_N + 6)] = '第三方'
    ws['C%d' % (inner_N + 6)] = data['trouble_time_out']
    ws['D%d' % (inner_N + 6)] = "%.2f%%" % ((float(data['trouble_time_out']) / data['trouble_time_all']) * 100)

    ws.merge_cells('A%d:B%d' % (inner_N+7, inner_N+7))
    ws['A%d' % (inner_N +7)] = '合计'
    ws['C%d' % (inner_N + 7)] = data['trouble_time_all']
    ws['D%d' % (inner_N + 7)] = "%.2f%%" % ((float(data['trouble_time_all']) / data['trouble_time_all']) * 100)

    for i in "ABCD":
        for n in range(inner_N,inner_N+8):
            ws['%s%d' % (i, n)].fill = headFill
            ws['%s%d' % (i, n)].alignment = headtext
            ws['%s%d' % (i, n)].protection = protection
            ws['%s%d' % (i, n)].border = border
            ws['%s%d' % (i, n)].border = border
            ws['%s%d' % (i, n)].border = border
            ws['%s%d' % (i, n)].font = ft

    for i in "ABCDEFGHIJK":
        ws['%s%d' % (i,inner_N)].fill = titleFill


    ###############################表格四###################################

    branch_N = inner_N + 16

    ws.merge_cells('A%d:K%d' % (branch_N, branch_N))
    ws['A%d' % branch_N] = u'各部门故障信息统计'

    ws['A%d' % (branch_N + 1)] = u'归属'
    ws['B%d' % (branch_N + 1)] = u'负责人'
    ws['C%d' % (branch_N + 1)] = u'核心服务'
    ws['D%d' % (branch_N + 1)] = u'非核心服务'
    ws['E%d' % (branch_N + 1)] = u'小计'


    ws['A%d' % (branch_N + 2)] = '运维'
    ws['B%d' % (branch_N + 2)] = '钱建峰'
    ws['C%d' % (branch_N + 2)] = data['trouble_time_yw_core']
    ws['D%d' % (branch_N + 2)] = data['trouble_time_yw_ncore']
    ws['E%d' % (branch_N + 2)] = data['trouble_time_yw']

    ws['A%d' % (branch_N + 3)] = '业务开发'
    ws['B%d' % (branch_N + 3)] = '肖朋'
    ws['C%d' % (branch_N + 3)] = data['trouble_time_ywkf_core']
    ws['D%d' % (branch_N + 3)] = data['trouble_time_ywkf_ncore']
    ws['E%d' % (branch_N + 3)] = data['trouble_time_ywkf']

    ws['A%d' % (branch_N + 4)] = '基础开发'
    ws['B%d' % (branch_N + 4)] = '黄谦'
    ws['C%d' % (branch_N + 4)] = data['trouble_time_jckf_core']
    ws['D%d' % (branch_N + 4)] = data['trouble_time_jckf_ncore']
    ws['E%d' % (branch_N + 4)] = data['trouble_time_jckf']

    ws['A%d' % (branch_N + 5)] = '运营'
    ws['B%d' % (branch_N + 5)] = '运营'
    ws['C%d' % (branch_N + 5)] = data['trouble_time_yy_core']
    ws['D%d' % (branch_N + 5)] = data['trouble_time_yy_ncore']
    ws['E%d' % (branch_N + 5)] = data['trouble_time_yy']

    ws['A%d' % (branch_N + 6)] = '第三方'
    ws['B%d' % (branch_N + 6)] = '第三方'
    ws['C%d' % (branch_N + 6)] = data['trouble_time_dsf_core']
    ws['D%d' % (branch_N + 6)] = data['trouble_time_dsf_ncore']
    ws['E%d' % (branch_N + 6)] = data['trouble_time_dsf']

    ws.merge_cells('A%d:B%d' % (branch_N + 7, branch_N + 7))
    ws['A%d' % (branch_N + 7)] = '合计'
    ws['C%d' % (branch_N + 7)] = data['trouble_time_core']
    ws['D%d' % (branch_N + 7)] = data['trouble_time_ncore']
    ws['E%d' % (branch_N + 7)] = data['trouble_time_all']

    for i in "ABCDE":
        for n in range(branch_N, branch_N + 8):
            ws['%s%d' % (i, n)].fill = headFill
            ws['%s%d' % (i, n)].alignment = headtext
            ws['%s%d' % (i, n)].protection = protection
            ws['%s%d' % (i, n)].border = border
            ws['%s%d' % (i, n)].border = border
            ws['%s%d' % (i, n)].border = border
            ws['%s%d' % (i, n)].font = ft
    for i in "ABCDEFGHIJK":
        ws['%s%d' % (i, branch_N)].fill = titleFill

    ###############################表格五###################################


    type_N= branch_N + 16

    ws.merge_cells('A%d:K%d' % (type_N, type_N))
    ws['A%d' % type_N] = '各类型故障信息统计'

    ws['A%d' % (type_N + 1)] = '故障类型'
    ws['B%d' % (type_N + 1)] = '故障时间'
    ws['C%d' % (type_N + 1)] = '占比'

    ws['A%d' % (type_N + 2)] = '服务器类型故障'
    ws['B%d' % (type_N + 2)] = data['trouble_time_server']
    ws['C%d' % (type_N + 2)] = "%.2f%%" % ((float(data['trouble_time_server']) / data['trouble_time_all']) * 100)

    ws['A%d' % (type_N + 3)] = '人为错误'
    ws['B%d' % (type_N + 3)] = data['trouble_time_perple']
    ws['C%d' % (type_N + 3)] = "%.2f%%" % ((float(data['trouble_time_perple']) / data['trouble_time_all']) * 100)

    ws['A%d' % (type_N + 4)] = 'BUG类型故障'
    ws['B%d' % (type_N + 4)] = data['trouble_time_bug']
    ws['C%d' % (type_N + 4)] = "%.2f%%" % ((float(data['trouble_time_bug']) / data['trouble_time_all']) * 100)

    ws['A%d' % (type_N + 5)] = '安全类型故障'
    ws['B%d' % (type_N + 5)] = data['trouble_time_safe']
    ws['C%d' % (type_N + 5)] = "%.2f%%" % ((float(data['trouble_time_safe']) / data['trouble_time_all']) * 100)

    ws['A%d' % (type_N + 6)] = '偶然性故障'
    ws['B%d' % (type_N + 6)] = data['trouble_time_once']
    ws['C%d' % (type_N + 6)] = "%.2f%%" % ((float(data['trouble_time_once']) / data['trouble_time_all']) * 100)

    ws['A%d' % (type_N + 7)] = '网络故障'
    ws['B%d' % (type_N + 7)] = data['trouble_time_net']
    ws['C%d' % (type_N + 7)] = "%.2f%%" % ((float(data['trouble_time_net']) / data['trouble_time_all']) * 100)

    ws['A%d' % (type_N + 8)] = '第三方故障'
    ws['B%d' % (type_N + 8)] = data['trouble_time_dsf_t']
    ws['C%d' % (type_N + 8)] = "%.2f%%" % ((float(data['trouble_time_dsf_t']) / data['trouble_time_all']) * 100)

    ws['A%d' % (type_N + 9)] = '合计'
    ws['B%d' % (type_N + 9)] = data['trouble_time_all']
    ws['C%d' % (type_N + 9)] = "%d%%" % 100


    for i in "ABC":
        for n in range(type_N,type_N+10):
            ws['%s%d' % (i,n)].fill = headFill
            ws['%s%d' % (i, n)].alignment = headtext
            ws['%s%d' % (i, n)].protection = protection
            ws['%s%d' % (i, n)].border = border
            ws['%s%d' % (i, n)].border = border
            ws['%s%d' % (i, n)].border = border
            ws['%s%d' % (i, n)].font = ft
    for i in "ABCDEFGHIJK":
        ws['%s%d' % (i, type_N)].fill = titleFill


    #画饼图
    pie = PieChart()
    labels = Reference(ws, min_col=2, min_row=inner_N + 2, max_row=inner_N + 6)
    data_1 = Reference(ws, min_col=3, min_row=inner_N + 2, max_row=inner_N + 6)
    pie.add_data(data_1, titles_from_data=False)
    pie.set_categories(labels)
    pie.title = "内外故障信息统计"

    ws.add_chart(pie, "G%d" % (inner_N + 1))


    pie = PieChart()
    labels = Reference(ws, min_col=2, min_row=branch_N + 2, max_row=branch_N + 6)
    data_1 = Reference(ws, min_col=5, min_row=branch_N + 2, max_row=branch_N + 6)
    pie.add_data(data_1, titles_from_data=False)
    pie.set_categories(labels)
    pie.title = "各部门故障信息统计"

    ws.add_chart(pie, "G%d" % (branch_N + 1))


    pie = PieChart()
    labels = Reference(ws, min_col=1, min_row=type_N+2, max_row=type_N+8)
    data_1 = Reference(ws, min_col=2, min_row=type_N+2, max_row=type_N+8)
    pie.add_data(data_1, titles_from_data=False)
    pie.set_categories(labels)
    pie.title = "各类型故障信息统计"

    ws.add_chart(pie, "G%d" % (type_N+1))






###################################################################################

    ws = wb.create_sheet(u'故障详细列表', 1)

    # 设置行高
    ws.row_dimensions[1].height = 40.0
    ws.row_dimensions[2].height = 40.0

    # ws['A1']=head


    trouble_infos = Trouble_repo.query.filter(Trouble_repo.trouble_date.ilike("%s%%" % data['this_month']),Trouble_repo.trouble_status==u'完成').order_by(Trouble_repo.trouble_date)

    trouble_list = []
    trouble_times_core = 0
    trouble_times_ncore = 0
    for i in trouble_infos:
        if i.isnot_core == '是':
            try:
                affect_time = int(i.affect_time)
            except:
                affect_time = 0
            trouble_times_core += affect_time
        else:
            try:
                affect_time = int(i.affect_time)
            except:
                affect_time = 0
            trouble_times_ncore += affect_time

        List = [i.trouble_date, i.operating_center, i.business_module, i.trouble_affair, i.affect_scope, i.isnot_inner,
                i.affect_time, i.isnot_experience, i.affect_user, i.affect_money,
                i.data_source, i.isnot_core, i.trouble_type, i.heading_user, i.trouble_attr, i.trouble_status,
                i.trouble_cause, i.whith_process, i.lesson_course, i.improve]

        trouble_list.append(List)

    stab_per_core = 0
    stab_per_ncore = 0

    stab_per_core = "%.2f" % (100 - float(trouble_times_core)/data['month_time']*100)
    stab_per_ncore = "%.2f" % (100 - float(trouble_times_ncore) / data['month_time'] * 100)

    ws['B1'] = u'核心服务故障时间:%s分钟 ' % trouble_times_core
    ws['C1'] = u'稳定性:%s%%' % stab_per_core
    ws['F1'] = u'非核心服务故障时间:%s分钟 ' % trouble_times_ncore
    ws['G1'] = u'稳定性:%s%%' % stab_per_ncore

    head_text = [u"日期", u"运营中心", u"业务模块", u"事件", u"影响范围", u"是否内部故障", u"影响时长(分钟)", u"是否影响用户体验", u"影响用户", u"直接经济损失(美元)",
                 u"数据来源", u"是否核心服务", u"故障类型", u"处理负责人", u"归属", u"状态", u"故障原因", u"处理过程", u"教训总结", u"改进"]

    ws.append(head_text)

    ws['C2'].comment = Comment(text="游戏明\n储值\n登陆\n后台\n所有", author="业务模块")
    ws['D2'].comment = Comment(text="描素事件现象", author="事件")
    ws['F2'].comment = Comment(text="如网络、CDN、原厂、渠道等外部因素引起的故障都不是内部故障", author="是否内部故障")
    ws['I2'].comment = Comment(text="1、昨天在线数据-今天储值数据；\n2、如数据值为负，表示故障时段数据比昨天上升，则故障影响不大；\n3、若数据为0，则表示无影响。", author="影响用户")
    ws['J2'].comment = Comment(text="1、昨天储值数据-今天储值数据；\n2、如数据值为负，表示故障时段数据比昨天上升，则故障影响不大 \n3、若数据为0，则表示无影响", author="经济损失")

    n = 3
    for trouble in trouble_list:

        ws.append(trouble)
        for i in 'ABCDEFGHIJKLMNOPQRST':
            ws['%s%d' % (i, n)].border = border
            ws['%s%d' % (i, n)].alignment = bodytext
            ws['%s%d' % (i, n)].protection = protection
        ws['Q%d' % n].alignment = Alignment(horizontal='general', vertical='center', wrap_text=True)
        ws['R%d' % n].alignment = Alignment(horizontal='general', vertical='center', wrap_text=True)
        ws['S%d' % n].alignment = Alignment(horizontal='general', vertical='center', wrap_text=True)
        ws['T%d' % n].alignment = Alignment(horizontal='general', vertical='center', wrap_text=True)
        n += 1

    for i in 'ABCDEFGHIJKLMNOPQRST':
        ws['%s1' % i].fill = titleFill
        ws['%s2' % i].fill = headFill

        ws['%s2' % i].border = border

        ws['%s1' % i].font = ft
        ws['%s2' % i].font = ft

        ws['%s1' % i].alignment = headtext
        ws['%s2' % i].alignment = headtext

        ws['%s1' % i].protection = protection

        # 设置宽度
        ws.column_dimensions[i].width = 30.0
    ws.column_dimensions['I'].width = 32.0

    ws['F2'].fill = PatternFill(start_color='e26b0a', end_color='e26b0a', fill_type='solid')
    ws['G2'].fill = PatternFill(start_color='e26b0a', end_color='e26b0a', fill_type='solid')

    # 正式环境
    # title = u'/opt/Flask/app/static/file/%s月故障分析.xlsx' % data['this_month']
    # 测试环境
    title = u'%s/app/static/files/report/%s月故障分析.xlsx' % (basedir,data['this_month'])

    wb.save(title)




